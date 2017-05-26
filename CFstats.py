import openpyxl

wb = openpyxl.load_workbook('CF stats.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

for i in range(1, sheet.get_highest_row()+1 ) :

	c=sheet.cell(row=i, column=1).value
	flag1=1
	flag2=0
	num1=0
	num2=0

	for k in range ( 1 , len(c)-1 ) :
	
		if ( c[k-1]==' ' and c[k]==' ' and c[k+1]!=' ' ):
			flag1=flag1+1
			num1=k+1
			print (k,'flag1',flag1)

		if ( c[k-1]!=' ' and c[k]==' ' and c[k+1]==' '  ):
			flag2=flag2+1
			num2=k
			print (k, 'flag2', flag2)

		if ( k==len(c)-2):
			num2=k+2	
		
		if(flag1>0 & ( flag1==flag2 or k==len(c)-2 ) ):
			sheet.cell(row=i,column=flag1+1).value=c[num1:num2]
			print ( c[num1:num2])



wb.save('b.xlsx')



			
	
